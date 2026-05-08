"""입결(입출력 결과) 데이터 표준 정의 페이지"""
from __future__ import annotations
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from lib.std import INPUT_STANDARD_SPEC, STAGES, CHK_HEADER, GRID_HEADER

st.set_page_config(page_title="데이터 표준", page_icon="📥", layout="wide")
st.markdown("# 📥 입결(입출력 결과) 데이터 표준")
st.caption(
    "이 대시보드는 두 가지 엑셀 파일을 표준 입력으로 받습니다. "
    "아래 명세대로 작성된 파일이라면 어떤 현장에도 동일하게 적용됩니다."
)

# ── 표준 명세 ────────────────────────────────────────────────
st.markdown("## 📑 표준 컬럼 명세")
spec_df = pd.DataFrame(INPUT_STANDARD_SPEC, columns=["필드", "한글명", "타입", "필수", "설명"])
spec_df["필수"] = spec_df["필수"].map({True: "✅", False: "—"})
st.dataframe(spec_df, use_container_width=True, hide_index=True)

# ── 두 파일 구조 설명 ─────────────────────────────────────────
st.markdown("## 📦 파일 ① — CHK LIST (안산 설치물량.xlsx)")
st.markdown(
    """
- **시트명**: `CHK LIST`
- **데이터 시작 행**: 8 (R1~R7은 헤더)
- **부재 단위 마스터** + **7단계 공정 진척** + **현장 시공 실적**을 한 행에 담는 입결 표준 시트
- 핵심 식별자: **부재명(member_no)** — 컬럼 P (16번째)
- 공정 단계마다 **계획일 / 실적일 / 수량 / 중량** 4개 컬럼 묶음 (전처리는 3개)
"""
)

stage_cols = [
    {"단계": "취부",   "계획": "X", "실적": "Y", "수량": "Z",  "중량": "AA"},
    {"단계": "용접",   "계획": "AB", "실적": "AC", "수량": "AD", "중량": "AE"},
    {"단계": "마감",   "계획": "AF", "실적": "AG", "수량": "AH", "중량": "AI"},
    {"단계": "검사",   "계획": "AJ", "실적": "AK", "수량": "AL", "중량": "AM"},
    {"단계": "전처리", "계획": "—",  "실적": "AN", "수량": "AO", "중량": "AP"},
    {"단계": "도장",   "계획": "AQ", "실적": "AR", "수량": "AS", "면적": "AT"},
    {"단계": "발송",   "계획": "AU", "실적": "AV", "수량": "AW", "중량": "AX"},
    {"단계": "시공",   "계획": "—",  "실적": "CL", "수량": "CM", "중량": "CN"},
]
st.dataframe(pd.DataFrame(stage_cols), use_container_width=True, hide_index=True)

st.markdown("## 📦 파일 ② — 그리드 마스터 (그리드정리.xlsx)")
st.markdown(
    """
- **시트명**: `Sheet1`
- **데이터 시작 행**: 2 (R1은 헤더)
- 부재명(DWGNO.)과 **GRID 위치 문자열**(`3-5/G-H` 형식)의 1:1 매핑 테이블
- CHK LIST의 부재명과 join 되어 시공 위치 시각화에 사용됨
"""
)
grid_cols = pd.DataFrame([
    {"컬럼": "A", "필드": "DWGNO.", "필수": "✅", "설명": "부재명 — CHK LIST와 조인 키"},
    {"컬럼": "B", "필드": "DESCRIPTION", "필수": "—", "설명": "부재 종류 (BEAM/COLUMN…)"},
    {"컬럼": "C", "필드": "MEMBERSIZE", "필수": "—", "설명": "단면 규격"},
    {"컬럼": "D", "필드": "Qty.", "필수": "—", "설명": "동일 부재 개수"},
    {"컬럼": "E", "필드": "Length", "필수": "—", "설명": "길이(mm)"},
    {"컬럼": "F", "필드": "TotalPaint", "필수": "—", "설명": "도장면적"},
    {"컬럼": "G", "필드": "Total", "필수": "—", "설명": "단위 중량"},
    {"컬럼": "H", "필드": "GIRD", "필수": "✅", "설명": "GRID 위치 문자열 (예: 3-5/G-H)"},
    {"컬럼": "I", "필드": "REMARK", "필수": "—", "설명": "비고"},
])
st.dataframe(grid_cols, use_container_width=True, hide_index=True)

# ── 빈 양식 다운로드 ──────────────────────────────────────────
st.markdown("---")
st.markdown("## 💾 빈 표준 양식 다운로드")


def _build_empty_chk() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CHK LIST"

    header_fill = PatternFill("solid", fgColor="1F3A68")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="맑은 고딕")
    border = Border(*[Side(style="thin", color="BBBBBB")] * 4)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 헤더 (R7 기준)
    headers = [
        "PKG NO.", "팀", "구분", "재질", "ITEM", "내역", "E-DWG", "절주",
        "X", "Y", "항차", "사이즈", "높이", "폭", "종류", "부재명",
        "MEMBERSIZE", "길이", "수량", "중량(kg)", "소수점환산", "값붙여넣기", "도장(M2)",
    ]
    # 단계별 4개 컬럼
    for stg in ["취부", "용접", "마감", "검사"]:
        headers += [f"{stg}(계)", f"{stg}(실)", f"{stg}(수)", f"{stg}(중)"]
    headers += ["전처리(실)", "전처리(수)", "전처리(중)"]
    headers += ["도장(계)", "도장(실)", "도장(수)", "도장(면적)"]
    headers += ["발송(계)", "발송(실)", "발송(수)", "발송(중)"]
    headers += ["송장번호"]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = align
        ws.column_dimensions[c.column_letter].width = max(10, min(18, len(str(h)) + 2))

    # 시공 컬럼 (C90~92)
    for col, h in [(90, "현장 일자"), (91, "현장 수량"), (92, "현장 중량"), (95, "GRID")]:
        c = ws.cell(row=7, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="F5A623")
        c.font = header_font
        c.border = border
        c.alignment = align
        ws.column_dimensions[c.column_letter].width = 14

    ws.row_dimensions[7].height = 38
    ws.freeze_panes = "Q8"  # 부재명 컬럼까지 고정

    # 안내 행
    ws.cell(row=2, column=1, value="◎ 입결 데이터 표준 양식 — 부재 1행 단위로 R8부터 입력")
    ws.cell(row=2, column=1).font = Font(bold=True, size=11, color="1F3A68")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_empty_grid() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_fill = PatternFill("solid", fgColor="1F3A68")
    header_font = Font(color="FFFFFF", bold=True, name="맑은 고딕")

    headers = ["DWGNO.", "DESCRIPTION", "MEMBERSIZE", "Qty.", "Length",
               "TotalPaint", "Total", "GIRD", "REMARK"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = 14

    ws.cell(row=2, column=1, value="A1B-001")
    ws.cell(row=2, column=2, value="BEAM")
    ws.cell(row=2, column=3, value="L50*50*4")
    ws.cell(row=2, column=4, value=1)
    ws.cell(row=2, column=5, value=9569)
    ws.cell(row=2, column=8, value="3-5/G-H")
    ws.freeze_panes = "B2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


c1, c2 = st.columns(2)
with c1:
    st.download_button(
        "📄 ① CHK LIST 빈 양식 다운로드",
        data=_build_empty_chk(),
        file_name="CHK_LIST_표준양식.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with c2:
    st.download_button(
        "📄 ② 그리드 마스터 빈 양식 다운로드",
        data=_build_empty_grid(),
        file_name="GRID_마스터_표준양식.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── 현재 로드된 데이터 다운로드 ────────────────────────────────
st.markdown("---")
st.markdown("## 📤 현재 로드된 통합 데이터 (CSV)")

from lib import get_dataset

df = get_dataset()
csv = df.to_csv(index=False).encode("utf-8-sig")
st.download_button(
    f"📥 통합 데이터 CSV 다운로드 ({len(df):,}행)",
    data=csv, file_name="안산_시공실적_통합.csv", mime="text/csv",
    use_container_width=True,
)
with st.expander("🔍 통합 데이터 미리보기 (상위 50행)", expanded=False):
    st.dataframe(df.head(50), use_container_width=True, hide_index=True)
